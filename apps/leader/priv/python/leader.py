import openpyxl
import emailer
import webSearcher
import os

# todo
# fill in missing info

# possibly send emails for each person


# abstractions
# take colomn names from raw
# use colomn names for sorting instead of a fixed order
# create modules for each task

# creates the new workbook
def save_wb(east_excel, west_excel, middle_excel):
    os.chdir('/Users/mechimcook/dev/Python-fun/leader/output/')
    wb = openpyxl.Workbook()


    west_leads = wb.create_sheet("west")
    west_leads.append(("CODE", "SLSMN", "Company name", "First name", "Last name","Phone Number","Email","Street Address","Town","Zip Code","State","Comments"))
    for row in west_excel:
        west_leads.append(tuple(row))

    east_leads = wb.create_sheet("east")
    east_leads.append(("CODE", "SLSMN", "Company name", "First name", "Last name","Phone Number","Email","Street Address","Town","Zip Code","State","Comments"))
    for row in east_excel:
        east_leads.append(tuple(row))

    middle_leads = wb.create_sheet("unsorted-middle")
    middle_leads.append(("CODE", "SLSMN", "Company name", "First name", "Last name","Phone Number","Email","Street Address","Town","Zip Code","State","Comments"))
    for row in middle_excel:
        middle_leads.append(tuple(row))
    wb.save('leads.xlsx')



# reads raw
workbook = openpyxl.load_workbook('White.xlsx')
raw = workbook.active
## TODO:
  # move states to new file for readability

east_states =['ME','MAINE','NH','NEW HAMPSHIRE','VT','VERMONT','NY','NEW YORK','MA',
'MASSACHUSETTS','RI','RHODE ISLAND','CT','CONNECTICUT','NJ','NEW JERSEY','PA', 'PENNSYLVANIA','DE',
'DELAWARE','MD','MARYLAND','DC','DISTRICT OF COLUMBIA', 'MI','MICHIGAN','OH','OHIO','IN','INDIANA','IL','ILLINOIS',
'WI','WISCONSIN','WV','WEST VIRGINIA' ,'VA','VIRGINIA','NC','NORTH CAROLINA' ,'TN','TENNESSEE','KY','KENTUCKY','SC',
'SOUTH CAROLINA','GA','GEORGIA','AL','ALABAMA','MS','MISSISSIPPI','FL','FLORIDA']
west_states =['WYOMING','WY',
'COLORADO','CO',
'UTAH','UT',
'NEVADA','NV',
'IDAHO','ID',
'CALIFORNIA','CA',
'OREGON','OR',
'WASHINGTON','WA',
'ALASKA','AK',
'MONTANA','MT']
east_excel = []
west_excel = []
middle_excel = []


# create email folder
folder_name = "/output/emails"
cwd = os.getcwd()
path = cwd + folder_name

try:
    os.makedirs(path)
except OSError:
    print ("Creation of the directory %s failed" % path)
else:
    print ("Successfully created the directory %s " % path)



for row in raw.iter_rows(min_row=2, values_only=True):
    # check for missing information d[0] to k[10]. a amd b require db conection
    lead = []
    for value in row:
      lead.append(value)
    lead = webSearcher.get_missing(lead)
    print(lead[10].upper()+" this")

    if lead[10].upper() in east_states:
        east_excel.append(lead)
    elif lead[10].upper() in west_states:
        west_excel.append(lead)
        emailer.email_west(lead)
    else:
        middle_excel.append(lead)
    save_wb(east_excel, west_excel, middle_excel)
